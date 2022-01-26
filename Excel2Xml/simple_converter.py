"""
Convertidor de Excel a XML
========================

Este programa convierte un archivo Excel generado por el módulo RTE a XML,
requerido por la Unidad de Análisis Financiera(UAF), a formato XML.
"""
import sys
from abc import ABCMeta
from time import time
from datetime import datetime

from openpyxl import load_workbook
from yattag import Doc, indent
from xml.etree.ElementTree import parse, ParseError
import pandas as pd
import PySimpleGUI as sg
# from resources import get_resources_path

SUCURSALES = {
    "0": 'PRINCIPAL',
    "1": 'SANTO DOMINGO',
    "2": 'HIGUEY',
    "3": 'SAN CRISTOBAL',
    "4": 'BARAHONA',
    "5": 'SAN JUAN DE LA MAGUANA',
    "6": 'SAN FRANCISCO MACORIS',
    "7": 'COMENDADOR',
    "8": 'COTUI',
    "9": 'LA VEGA',
    "10": 'SANTIAGO RODRIGUEZ',
    "11": 'MONTE CRISTI',
    "12": 'PUERTO PLATA',
    "13": 'NAGUA',
    "15": 'EL SEYBO',
    "16": 'SANTIAGO',
    "17": 'SAN JOSE DE OCOA',
    "18": 'AZUA',
    "19": 'BANI',
    "20": 'VALVERDE MAO',
    "21": 'ARENOSO',
    "22": 'HATO MAYOR',
    "23": 'MOCA',
    "24": 'SAMANA',
    "25": 'BONAO',
    "26": 'NEYBA',
    "27": 'DAJABON',
    "28": 'SAN JOSE DE LAS MATAS',
    "29": 'RIO SAN JUAN',
    "30": 'VILLA RIVA',
    "31": 'SALCEDO',
    "32": 'MONTE PLATA',
    "33": 'CONSTANZA',
}

DIRECCIONES = {
    "PRINCIPAL": "Ave.George Washington NO. 601",
    "SANTO DOMINGO": "Ave.George Washington NO. 601",
    "SANTIAGO RODRIGUEZ": "C/ Dr. Darío Gómez No. 64",
    "MONTE CRISTI": "C/ BENITO MONCION, ESQ. SANCHEZ NO.60",
    "PUERTO PLATA": "C/ PRINCIPAL PROF. JUAN BOSCH NO.4",
    "CONSTANZA": "AV. ANTONIO ABUD ISSAC",
    "NAGUA": "C/ 27 DE FEBRERO, ESQ. MERCEDES BELLO NO.24",
    "EL SEYBO": "C/ MANUELA DIEZ JIMENEZ No.10",
    "SANTIAGO": "AV. JUAN PABLO DUARTE, ESQ. ESTADO DE ISRAEL",
    "SAN JOSE DE OCOA": "C/ DUARTE ESQ. ALTAGRACIA No.40",
    "AZUA": "C/ EMILIO PRUDHOMME No.35",
    "BANI": "C/ NUESTRA SRA. DE REGLA No.19 ESQ. SANCHEZ",
    "HIGUEY": "C/ ALTAGRACIA ESQ. LAS CARRERAS",
    "VALVERDE MAO": "AV. BENITO MONCION",
    "ARENOSO": "AV. DUARTE No.61",
    "HATO MAYOR": "C/ MELCHOR CONTIN ALPHAU No.39, MILLON",
    "MOCA": "CARR. MOCA LA VEGA, AL LADO DEL HOSPITAL",
    "SAMANA": "C/ SANTA BARBARA, EDIF. OFICINAS PUBLICAS",
    "BONAO": "C/ DUARTE No.279, SECTOR EL 90",
    "NEYBA": "C/ SAN BARTOLOME No.41",
    "DAJABON": "AV. PABLO REYES NO.29",
    "SAN JOSE DE LAS MATAS": "C/ 16 DE AGOSTO No.18, CENTRO DEL PUEBLO",
    "RIO SAN JUAN": "AUTOPISTA DR. DOMINGO ANTIONIO GONZALEZ",
    "SAN CRISTOBAL": "C/ GENERAL CABRAL No. 34",
    "VILLA RIVA": "C/ 27 DE FEBRERO ESQ. COLON NO.57",
    "MONTE PLATA": "C/ GENERAL MATIAS MORENO ESQ. ALTAGRACIA No. 7",
    "BARAHONA": "AV. LUIS E. DEL MONTE No.59, CENTRO DE LA CIU",
    "SAN JUAN DE LA MAGUANA": "C/ SANCHEZ No.67",
    "SAN FRANCISCO MACORIS": "Av. Frank Grullón, Salida hacia Nagua",
    "SALCEDO": "C/ DUARTE, ESQ. HERMANAS MIRABAL",
    "COMENDADOR": "C/ 27 DE FEB. ESQ. LUZ CELESTE LARA No.20",
    "COTUI": "C/ PADRE FANTINO NO.5",
    "LA VEGA": "PROFESOR JUAN BOSCH, ESQ. COMANDANTE JIMÉNEZ",
}

# RTEMAP = get_resources_path("data/mapped_elements.xlsx")


class IXMLFormatter(metaclass=ABCMeta):
    "Class interface for (Converter)"
    @classmethod
    def __get_wb(cls):
        "Interface method that intends to create instance object"

    @staticmethod
    @classmethod
    def get_contact(cls, key):
        "Interface method intends to get data from a different source"

    @staticmethod
    @classmethod
    def gen_xml(cls):
        "Interface method that intends to generate the concrete xml file"

    @staticmethod
    @classmethod
    def save_obj(cls, data):
        """Interface method that intends to save the xml object as a file"""

    @staticmethod
    @classmethod
    def success_msg(cls):
        """Interface method that intends to create a success message window popup"""


class XMLFormatter(IXMLFormatter):
    "Concrete class that implements the XMLFormatter interface"
    def __init__(self):
        self.name = "RTEXML"
        self.wb = None
        self.__get_wb()

    def __repr__(self) -> str:
        return self.name

    def __len__(self) -> int:
        df = pd.read_excel(self.wb, sheet_name='V2.5')
        return len(df.index)

    def __get_wb(self) -> None:
        """Load Excel file to convert"""
        self.wb = sys.argv[1] if len(sys.argv) > 1 else sg.popup_get_file("Seleccionar archivo...")
        if not self.wb:
            sg.popup("Cancelando...", "No se seleccionó ningún archivo")
            raise SystemExit("Proceso cancelado")

    def success_msg(self) -> None:
        """Message successful operation"""
        sg.popup("Archivo XML generado exitosamente.", self.wb)

    def get_contact(self, key: str) -> None:
        """List of entities which are missing in RTE program from Banke module"""
        df = pd.read_excel('representantes_de_entidades.xlsx', sheet_name='Entidades')

        table = df[
            ['RNC', 'CONTACTO', 'APELLIDO', 'IDENTIFICACION', 'SEXO', 'FECHA NACIMIENTO', 'TELEFONO', 'DIRECCION',
             'NACIONALIDAD', 'OCUPACION']]

        # FIXME There is a better way to loop
        for ind, reg in table.iterrows():
            if reg['RNC'].strip() == key.strip():
                return reg[
                    ['CONTACTO', 'APELLIDO', 'IDENTIFICACION', 'SEXO', 'FECHA NACIMIENTO', 'TELEFONO', 'DIRECCION',
                     'NACIONALIDAD', 'OCUPACION']]

    def save_obj(self, data: str) -> None:
        """Save string to a .xml file."""
        # ? Se agrega la fecha del día en que se genera el XML
        date = ''.join(str(datetime.now()).replace('-', '_')[:10])
        filename = "_UAF_Web_Report_" + date + ".xml"
        # Saving to file
        with open(filename, "w") as f:
            f.write(data)

        # ? Once generated, edit file to remove root node attribs.
        try:
            tree = parse(filename)
            if tree:
                tree.getroot().attrib.popitem()
                tree.write(filename)
        except ParseError as e:
            print('>>Exception<<: File not well-formed')

    def gen_xml(self) -> None:
        """
        :param workbook: str - Archivo Excel a convertir.
        :return: void
        """
        # Load our Excel File
        wb = load_workbook(self.wb)
        # Getting an object of active sheet 1
        ws = wb.worksheets[0]
        # Returning returns a triplet
        doc, tag, text = Doc().tagtext()

        xml_header = '<?xml version="1.0" encoding="iso-8859-3"?>'
        # xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema-instance"></xs:schema>'

        # ? Agrega el header al documento.
        doc.asis(xml_header)

        # ? Se crean los elementos o nodos.
        # ? El nodo report(root) se debe generar con el esquema especificado para
        # ? que el reporte siga las reglas definidas en la plataforma goAML.
        with tag('report', ('xmlns:xsi', "http://www.w3.org/2001/XMLSchema-instance"),
                 ('xsi:noNamespaceSchemaLocation', "goAMLSchema.xsd")):
            # ? Se recorre solo la primera fila para crear el esqueleto
            # ? Algunos datos están hardcode ya que el RTE no genera esta data o
            # ? estos datos serán siempre fijos.

            for row in ws.iter_rows(min_row=8, max_row=8, min_col=1, max_col=ws.max_column):
                # generate rows on demand
                row = [cell.value for cell in row if row is not None]
                # ? La indentación define la relación padre-hijo entre los nodos.
                with tag('rentity_id'):
                    text(int('1059'))  # int
                with tag('rentity_branch'):
                    text("Principal")  # str
                with tag('submission_code'):
                    if row[36] == "EFECTIVO" or row[36] == "CHEQUE":
                        text("E")
                    else:
                        text("E")  # Si existe otro método cambiar letra según documento.
                with tag('report_code'):
                    text("CTR")
                with tag('submission_date'):
                    tmp = datetime.strptime(str(row[4]), "%Y-%m-%d %H:%M:%S")
                    text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))  # Formato válido "%Y-%m-%dT%H:%M:%S"
                with tag("currency_code_local"):
                    text("DOP")
                with tag("reporting_person"):
                    with tag("gender"):
                        text("F")
                    with tag("title"):
                        text("Oficial de Cumplimiento")
                    with tag("first_name"):
                        text("Naty")
                    with tag("last_name"):
                        text("Abreu")
                    with tag("birthdate"):
                        tmp = datetime.strptime("1972-10-03 00:00:00", "%Y-%m-%d %H:%M:%S")
                        text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                    with tag("id_number"):
                        text("00109551382")
                    with tag("nationality1"):
                        text("DO")
                    with tag("phones"):
                        with tag("phone"):
                            with tag("tph_contact_type"):
                                text(2)
                            with tag("tph_communication_type"):
                                text("L")
                            with tag("tph_country_prefix"):
                                text("1")
                            with tag("tph_number"):
                                text("8095358088")
                            with tag("tph_extension"):
                                text("3212")
                    with tag("addresses"):
                        with tag("address"):
                            with tag("address_type"):
                                text(int("2"))
                            with tag("address"):
                                text("Ave. George Washington NO. 601")
                            with tag("town"):
                                text("Santo Domingo")
                            with tag("city"):
                                text("Santo Domingo")
                            with tag("zip"):
                                text("10103")
                            with tag("country_code"):
                                text("DO")
                    with tag("email"):
                        text("N.abreu@bagricola.gob.do")
                    with tag("occupation"):
                        text("Ing. de Sistemas")

                with tag("location"):
                    with tag("address_type"):
                        text(int("2"))
                    with tag("address"):
                        text("Calle Ave. George Washington No. 601")
                    with tag("city"):
                        text("SANTO DOMINGO,D.N.")
                    with tag("country_code"):
                        text("DO")
                    with tag("state"):
                        text("SANTO DOMINGO")

                with tag("reason"):
                    if row[74] is None or row[74] == '':
                        row[74] = "Transacciones que sobrepasaron los USD$15,000.00"
                        text(row[74])
                    else:
                        text(row[74])

            #? -------------------------------- TRANSACCIONS ---------------------------------#

            # Una vez que el esqueleto XML está hecho se recorren todas las filas
            # y se añade un nodo transacción por fila del RTE(Archivo Excel)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                row = [cell.value for cell in row if row is not None]
                # ? Si tiene número de reporte entonces la fila es válida
                if row[0]:
                    with tag("transaction"):
                        with tag("transactionnumber"):
                            text(row[77])
                        with tag("transaction_location"):
                            suc = DIRECCIONES.get(SUCURSALES.get(str(row[3])))
                            text(suc)
                        with tag("transaction_description"):
                            if row[41] is None:
                                row[41] = ""
                                text(row[41])
                            else:
                                text(row[41])
                        with tag("date_transaction"):
                            if row[37] is None:
                                row[37] = datetime.strptime("1900-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")
                            tmp = datetime.strptime(str(row[37]), "%Y-%m-%d %H:%M:%S")
                            text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                        with tag("teller"):
                            text(SUCURSALES.get(str(row[3])).upper())
                        with tag("authorized"):
                            if row[22] is None:
                                row[22] = SUCURSALES.get(str(row[3]))
                                text(row[22])
                            else:
                                text(row[22].strip())
                        with tag("transmode_code"):
                            text("6")
                        with tag("transmode_comment"):
                            if row[31] is None:
                                row[31] = ""
                                text(row[31])
                            else:
                                text(row[31].strip())
                        with tag("amount_local"):
                            if row[34] is None or row[34] == '':
                                row[34] = 0.0
                                text(row[34])
                            else:
                                text(row[34])
                        # ? SI EL BENEFICIARIO ES UNA ENTIDAD
                        con = get_contact(str(row[15]))

                        if row[6] == "JURIDICA":
                            if con is None:
                                print(f'Relacionado ')
                            with tag("t_from_my_client"):
                                with tag("from_funds_code"):
                                    text("K")
                                if row[48].lower() == 'no':
                                    with tag("t_conductor"):
                                        with tag("gender"):
                                            if con is None:
                                                text(row[82])
                                            else:
                                                text(con['SEXO'])
                                        with tag("title"):
                                            if con['SEXO'] == "M" or row[82] == "M":
                                                text("Sr")
                                            elif con['SEXO'] == "F" or row[82] == "F":
                                                text("Sra")
                                            else:
                                                text("n/a")
                                        with tag("first_name"):
                                            if con['CONTACTO'] is None or con['CONTACTO'] == '':
                                                con['CONTACTO'] = "n/a"
                                                text(con['CONTACTO'])
                                            # elif CONTACTS[row[15]]:
                                            #     text(''.join(CONTACTS[row[15]]))
                                            else:
                                                text(''.join(con['CONTACTO']))
                                        with tag("last_name"):
                                            if con['APELLIDO'] is None or con['APELLIDO'] == '' or con[
                                                'APELLIDO'] == '.':
                                                con['APELLIDO'] = "n/a"
                                                text(con['APELLIDO'])
                                            else:
                                                text(''.join(con['APELLIDO'].strip()))
                                        with tag("birthdate"):
                                            if con['FECHA NACIMIENTO'] is None:
                                                con['FECHA NACIMIENTO'] = datetime.strptime("1900-01-01 00:00:00",
                                                                                            "%Y-%m-%d %H:%M:%S")
                                            tmp = datetime.strptime(str(con['FECHA NACIMIENTO']), "%Y-%m-%d %H:%M:%S")
                                            text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                                        with tag("id_number"):
                                            if con['IDENTIFICACION'] is None:
                                                con['IDENTIFICACION'] = "n/a"
                                                text(con['IDENTIFICACION'])
                                            else:
                                                text(''.join(con['IDENTIFICACION'].split('-')).strip())
                                        with tag("nationality1"):
                                            text("DO")
                                        if row[14].lower() == "pasaporte":
                                            with tag("nationality2"):
                                                text("US")
                                        with tag("residence"):
                                            text("DO")
                                        with tag("phones"):
                                            with tag("phone"):
                                                with tag("tph_contact_type"):
                                                    text(2)
                                                with tag("tph_communication_type"):
                                                    text("L")
                                                with tag("tph_country_prefix"):
                                                    if con['TELEFONO'] is None:
                                                        con['TELEFONO'] = "n/a"
                                                        text(con['TELEFONO'])
                                                    else:
                                                        text(con['TELEFONO'].split('-')[0].strip())

                                                with tag("tph_number"):
                                                    if con['TELEFONO'] is None:
                                                        con['TELEFONO'] = "n/a"
                                                        text(con['TELEFONO'])
                                                    else:
                                                        text(''.join(con['TELEFONO'].split('-')[1].strip()))
                                        with tag("occupation"):
                                            if row[89] is None or row[89] == '':
                                                text("n/a")
                                            else:
                                                text(row[89])
                                        with tag("identification"):
                                            with tag("type"):
                                                text(1)
                                            with tag("number"):
                                                if row[84] is None or row[84] == '':
                                                    text("")
                                                else:
                                                    text(''.join(con['IDENTIFICACION'].split('-')).strip())
                                            with tag("issue_country"):
                                                if row[14].lower() == 'pasaporte':
                                                    text("US")
                                                else:
                                                    text("DO")

                                with tag("from_entity"):
                                    with tag("name"):
                                        if row[10] is None:
                                            row[10] = ""
                                            text(row[10])
                                        else:
                                            text(row[10].strip())
                                    with tag("commercial_name"):
                                        text(row[11])
                                    with tag("incorporation_legal_form"):
                                        text("C")
                                    with tag("incorporation_number"):
                                        if row[15] is None or row[15] == '':
                                            text("n/a")
                                        else:
                                            text(''.join(row[15].split('-')).strip())
                                    with tag("business"):
                                        text(row[17])
                                    with tag("phones"):
                                        with tag("phone"):
                                            with tag("tph_contact_type"):
                                                text(4)
                                            with tag("tph_communication_type"):
                                                text("M")
                                            with tag("tph_country_prefix"):
                                                text("1")
                                            with tag("tph_number"):
                                                if row[27] is None and row[28] is None and row[26] is None:
                                                    text("n/a")
                                                elif row[27]:
                                                    text(''.join(row[27].split('-')).strip())
                                                elif row[26]:
                                                    text(''.join(row[26].split('-')).strip())
                                                elif row[28]:
                                                    text(''.join(row[28].split('-')).strip())
                                    with tag("addresses"):
                                        with tag("address"):
                                            with tag("address_type"):
                                                text(int("4"))
                                            with tag("address"):
                                                if row[79] is None or row[79] == '':
                                                    text("n/a")
                                                else:
                                                    text(row[79].strip())
                                            with tag("city"):
                                                if row[23] is None or row[23] == '':
                                                    text("n/a")
                                                else:
                                                    text(row[23].strip())
                                            with tag("country_code"):
                                                text("DO")
                                            with tag("state"):
                                                if row[22] is None or row[22] == '':
                                                    text("n/a")
                                                else:
                                                    text(row[22].strip())

                                with tag("from_country"):
                                    text("DO")
                        # ? SI EL BENEFICIARIO ES UNA PERSONA
                        else:
                            with tag("t_from_my_client"):
                                with tag("from_funds_code"):
                                    text("K")
                                with tag("from_person"):
                                    with tag("gender"):
                                        if row[9] is None or row[9] == '':
                                            row[9] = "M"
                                            text(row[9])
                                        else:
                                            text(row[9])
                                    with tag("title"):
                                        if row[9] == "M":
                                            text("Sr")
                                        elif row[9] == "F":
                                            text("Sra")
                                        else:
                                            text("n/a")
                                    with tag("first_name"):
                                        text(row[10])
                                    with tag("last_name"):
                                        if row[11] is None or row[11] == '':
                                            row[11] = "."
                                            text(row[11])
                                        else:
                                            text(row[11])
                                    with tag("birthdate"):
                                        text("1988-06-17T00:00:00")
                                    with tag("id_number"):
                                        if row[14].lower() == "pasaporte":
                                            text(''.join(str(row[15][2:]).split('-')).strip())
                                        else:
                                            text(''.join(str(row[15]).split('-')).strip())
                                    with tag("nationality1"):
                                        text("DO")
                                    if row[14].lower() == "pasaporte":
                                        with tag("nationality2"):
                                            text("US")
                                    with tag("residence"):
                                        text("DO")
                                    with tag("phones"):
                                        with tag("phone"):
                                            with tag("tph_contact_type"):
                                                text(1)
                                            with tag("tph_communication_type"):
                                                if row[28] is None or row[28] == '':
                                                    text("L")
                                                else:
                                                    text("M")
                                            with tag("tph_country_prefix"):
                                                if row[27] is None and row[28] is None and row[26] is None:
                                                    text("n/a")
                                                elif row[27]:
                                                    text("1")
                                                elif row[28]:
                                                    text("1")
                                                elif row[26]:
                                                    text("1")
                                            with tag("tph_number"):
                                                if row[27] is None and row[28] is None and row[26] is None:
                                                    text("n/a")
                                                elif row[27]:
                                                    text(''.join(row[27].split('-')).strip())
                                                elif row[26]:
                                                    text(''.join(row[26].split('-')).strip())
                                                elif row[28]:
                                                    text(''.join(row[28].split('-')).strip())
                                    with tag("addresses"):
                                        with tag("address"):
                                            with tag("address_type"):
                                                text(int("1"))
                                            with tag("address"):
                                                if row[25] is None or row[25] == '' or row[25] == ' ':
                                                    text("No disponible")
                                                else:
                                                    text(''.join(row[25].split('/')))
                                            with tag("city"):
                                                if row[23] is None or row[23] == '':
                                                    text(SUCURSALES[str(row[3])])
                                                else:
                                                    text(row[23].strip())
                                            with tag("country_code"):
                                                text("DO")
                                            with tag("state"):
                                                if row[22] is None or row[22] == '':
                                                    text("n/a")
                                                else:
                                                    text(row[22].strip())
                                    with tag("occupation"):
                                        if row[17] is None or row[17] == '':
                                            text("n/a")
                                        else:
                                            text(row[17])
                                    with tag("identification"):
                                        with tag("type"):
                                            if row[14].lower() == "pasaporte":
                                                text("C")
                                            else:
                                                text(1)
                                        with tag("number"):
                                            if row[15] is None or row[15] == '':
                                                text("")
                                            elif row[14].lower() == "pasaporte":
                                                text(''.join(str(row[15][2:]).split('-')).strip())
                                            else:
                                                text(''.join(str(row[15]).split('-')).strip())
                                        with tag("issue_country"):
                                            if row[14].lower() == 'pasaporte':
                                                text("US")
                                            else:
                                                text("DO")
                                with tag("from_country"):
                                    text("DO")

                        # ? TAG REQUERIDO
                        with tag("t_to_my_client"):
                            with tag("to_funds_code"):
                                text("K")
                            with tag("to_account"):
                                with tag("institution_name"):
                                    text("Banco Agrícola de la República Dominicana")
                                with tag("swift"):
                                    text("401007665")
                                with tag("branch"):
                                    text(SUCURSALES.get(str(row[3])))
                                with tag("account"):
                                    text(row[19])
                                with tag("currency_code"):
                                    text("DOP")
                                with tag("account_name"):
                                    text(row[10])
                                with tag("client_number"):
                                    text(row[78])
                                with tag("personal_account_type"):
                                    text("C")
                                if row[6] == "JURIDICA":
                                    with tag("t_entity"):
                                        with tag("name"):
                                            text(row[10])
                                        with tag("commercial_name"):
                                            text(row[11])
                                        with tag("incorporation_number"):
                                            text(''.join(row[15].split('-')).strip())
                                        # with tag("incorporation_legal_form"):
                                        #     text("C")
                                        with tag("business"):
                                            text(row[17])
                                        with tag("phones"):
                                            with tag("phone"):
                                                with tag("tph_contact_type"):
                                                    text(4)
                                                with tag("tph_communication_type"):
                                                    text("M")
                                                with tag("tph_country_prefix"):
                                                    if row[27] is None and row[28] is None and row[26] is None:
                                                        text("n/a")
                                                    elif row[27]:
                                                        text("1")
                                                    elif row[28]:
                                                        text("1")
                                                    elif row[26]:
                                                        text("1")
                                                with tag("tph_number"):
                                                    if row[27] is None and row[28] is None and row[26] is None:
                                                        text("n/a")
                                                    elif row[27]:
                                                        text(''.join(row[27].split('-')).strip())
                                                    elif row[26]:
                                                        text(''.join(row[26].split('-')).strip())
                                                    elif row[28]:
                                                        text(''.join(row[28].split('-')).strip())
                                        with tag("addresses"):
                                            with tag("address"):
                                                with tag("address_type"):
                                                    text(int("4"))
                                                with tag("address"):
                                                    if row[79] is None or row[79] == '':
                                                        text("No Disponible")
                                                    else:
                                                        text(row[79].strip())
                                                with tag("city"):
                                                    if row[23] is None or row[23] == '':
                                                        text("n/a")
                                                    else:
                                                        text(row[23].strip())
                                                with tag("country_code"):
                                                    text("DO")
                                                with tag("state"):
                                                    if row[22] is None or row[22] == '':
                                                        text("n/a")
                                                    else:
                                                        text(row[22].strip())
                                        with tag("incorporation_country_code"):
                                            text("DO")
                                        with tag("incorporation_date"):
                                            if row[83] is None:
                                                row[83] = datetime.strptime("1945-08-29 00:00:00", "%Y-%m-%d %H:%M:%S")
                                            tmp = datetime.strptime(str(row[83]), "%Y-%m-%d %H:%M:%S")
                                            text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                                # ? Datos del representante
                                with tag("signatory"):
                                    with tag("is_primary"):
                                        text("true")
                                    with tag("t_person"):
                                        if row[6] == 'JURIDICA':
                                            with tag("gender"):
                                                if con['SEXO'] is None or con['SEXO'] == '':
                                                    con['SEXO'] = "M"
                                                    text(con['SEXO'])
                                                else:
                                                    text(con['SEXO'])
                                            with tag("first_name"):
                                                if con['CONTACTO'] is None or con['CONTACTO'] == '':
                                                    con['CONTACTO'] = "n/a"
                                                    text(con['CONTACTO'])
                                                else:
                                                    text(con['CONTACTO'].strip())
                                            with tag("last_name"):
                                                if con['APELLIDO'] is None or con['APELLIDO'] == '' or con[
                                                    'APELLIDO'] == '.':
                                                    con['APELLIDO'] = "n/a"
                                                    text(con['APELLIDO'])
                                                else:
                                                    text(con['APELLIDO'].strip())
                                            with tag("birthdate"):
                                                if con['FECHA NACIMIENTO'] is None:
                                                    con['FECHA NACIMIENTO'] = datetime.strptime("1900-01-01 00:00:00",
                                                                                                "%Y-%m-%d %H:%M:%S")
                                                    tmp = datetime.strptime(str(con['FECHA NACIMIENTO']),
                                                                            "%Y-%m-%d %H:%M:%S")
                                                    text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                                                else:
                                                    text("1988-06-17T00:00:00")
                                            with tag("id_number"):
                                                if con['IDENTIFICACION'] is None or con['IDENTIFICACION'] == '':
                                                    con['IDENTIFICACION'] = "n/a"
                                                    text(con['IDENTIFICACION'])
                                                else:
                                                    text(''.join(con['IDENTIFICACION'].split('-')).strip())
                                            with tag("nationality1"):
                                                text("DO")
                                            if row[14].lower() == "pasaporte":
                                                with tag("nationality2"):
                                                    text("US")
                                            with tag("residence"):
                                                text("DO")
                                            with tag("phones"):
                                                if row[6] == 'JURIDICA':
                                                    with tag("phone"):
                                                        with tag("tph_contact_type"):
                                                            text(1)
                                                        with tag("tph_communication_type"):
                                                            text("M")
                                                        with tag("tph_country_prefix"):
                                                            if con['TELEFONO'] is None:
                                                                text("n/a")
                                                            else:
                                                                text("1")
                                                        with tag("tph_number"):
                                                            if con['TELEFONO'] is None:
                                                                text("n/a")
                                                            else:
                                                                text(''.join(con['TELEFONO'].split('-')).strip())
                                                else:
                                                    with tag("phone"):
                                                        with tag("tph_contact_type"):
                                                            text(1)
                                                        with tag("tph_communication_type"):
                                                            text("M")
                                                        with tag("tph_country_prefix"):
                                                            if row[27] is None and row[28] is None and row[26] is None:
                                                                text("n/a")
                                                            elif row[27]:
                                                                text("1")
                                                            elif row[28]:
                                                                text("1")
                                                            elif row[26]:
                                                                text("1")
                                                        with tag("tph_number"):
                                                            if row[27] is None and row[28] is None and row[26] is None:
                                                                text("n/a")
                                                            elif row[27]:
                                                                text(''.join(row[27].split('-')).strip())
                                                            elif row[26]:
                                                                text(''.join(row[26].split('-')).strip())
                                                            elif row[28]:
                                                                text(''.join(row[28].split('-')).strip())

                                        else:
                                            with tag("gender"):
                                                if row[82] is None or row[82] == '':
                                                    row[82] = "M"
                                                    text(row[82])
                                                else:
                                                    text(row[82])

                                            with tag("first_name"):
                                                if row[80] is None or row[80] == '':
                                                    row[80] = "n/a"
                                                    text(row[80])
                                                else:
                                                    text(row[80].strip())

                                            with tag("last_name"):
                                                if row[81] is None or row[81] == '':
                                                    row[81] = "n/a"
                                                    text(row[81])
                                                else:
                                                    text(row[81].strip())

                                            with tag("birthdate"):
                                                if row[83] is None or row[83] == '':
                                                    row[83] = datetime.strptime("1900-01-01 00:00:00",
                                                                                "%Y-%m-%d %H:%M:%S")
                                                tmp = datetime.strptime(str(row[83]), "%Y-%m-%d %H:%M:%S")
                                                text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))

                                            with tag("id_number"):
                                                if row[84] is None or row[84] == '':
                                                    row[84] = "n/a"
                                                    text(row[84])
                                                elif row[14].lower() == "pasaporte":
                                                    text(''.join(str(row[84][2:]).split('-')).strip())
                                                else:
                                                    text(''.join(str(row[84]).split('-')).strip())

                                            with tag("nationality1"):
                                                text("DO")
                                            if row[14].lower() == "pasaporte":
                                                with tag("nationality2"):
                                                    text("US")
                                            with tag("residence"):
                                                text("DO")

                                            with tag("phones"):
                                                with tag("phone"):
                                                    with tag("tph_contact_type"):
                                                        text(1)
                                                    with tag("tph_communication_type"):
                                                        text("M")
                                                    with tag("tph_country_prefix"):
                                                        if row[27] is None and row[28] is None and row[26] is None:
                                                            text("n/a")
                                                        elif row[27]:
                                                            text("1")
                                                        elif row[28]:
                                                            text("1")
                                                        elif row[26]:
                                                            text("1")
                                                    with tag("tph_number"):
                                                        if row[85] is None or row[85] == '':
                                                            text("n/a")
                                                        else:
                                                            text(''.join(row[85].split('-')).strip())
                                            with tag("occupation"):
                                                if row[88] is None or row[88] == '':
                                                    text("n/a")
                                                else:
                                                    text(row[88].strip())
                                            with tag("identification"):
                                                with tag("type"):
                                                    if row[14].lower() == "pasaporte":
                                                        text("C")
                                                    else:
                                                        text("1")
                                                with tag("number"):
                                                    if row[84] is None or row[84] == '':
                                                        row[84] = "n/a"
                                                        text(row[84])
                                                    elif row[14].lower() == "pasaporte":
                                                        text(''.join(str(row[84][2:]).split('-')).strip())
                                                    else:
                                                        text(''.join(str(row[84]).split('-')).strip())
                                                with tag("issue_country"):
                                                    if row[14].lower() == 'pasaporte':
                                                        text("US")
                                                    else:
                                                        text("DO")
                                    with tag("role"):
                                        text("A")
                                with tag("opened"):  # FIXME: Fecha apertura cta
                                    if row[37] is None or row[37] == '':
                                        row[37] = datetime.strptime("1982-02-18 00:00:00", "%Y-%m-%d %H:%M:%S")
                                    tmp = datetime.strptime(str(row[37]), "%Y-%m-%d %H:%M:%S")
                                    text(tmp.strftime("%Y-%m-%dT%H:%M:%S"))
                                with tag("balance"):
                                    if row[34] is None or row[34] == '':
                                        row[34] = 0.0
                                        text(row[34])
                                    else:
                                        text(row[34])
                                with tag("status_code"):
                                    text("A")
                                with tag("beneficiary"):
                                    text(''.join(row[11]))

                            with tag("to_country"):
                                text("DO")


        # doc.getvalue() contiene el archivo XML en formato string
        result = indent(
            doc.getvalue(),
            indentation='   ',
            indent_text=False
        )

        self.save_obj(result)
        self.success_msg()


class ConverterFactory:
    "Factory class"
    @staticmethod
    def create_obj(prop):
        "Static method to generate a Converter Object"
        if prop == 'xml':
            return XMLFormatter()
        return None

#? --------------------------- main program class 1st version --------------------------------------#
# class Converter:
#     """This class converts an RTE module excel file to a valid
#         XML to meet goAML platform requirements.
#     """
#     # DEBUG time program execution
#     # start = time.time()
#
#     def __init__(self):
#         self.k = []
#         self.ele = []
#         self.wb = None
#
#     def run(self):
#         self.wb = sys.argv[1] if len(sys.argv) > 1 else sg.popup_get_file('Seleccionar archivo')
#         if not self.wb:
#             sg.popup("Cancelando", "No se seleccionó ningún archivo.")
#             raise SystemExit("Cancelado.")
#         else:
#             start = time()
#             gen_xml(self.wb)
#             sg.popup("Archivo XML generado exitosamente.", self.wb)
#             print(f'Execution time: {time() - start}')
#     # time.sleep(1)
#     # stop = time.time()
#     # print(f"Program execution with time module: {stop - start}")
#? --------------------------------------------------------------------------------------------------#


#? ------------ Utility functions ------------#
def get_contact(key: str):
    """List of entities which are missing in RTE program from Banke module"""
    # Easier to manipulate excel data with pandas
    df = pd.read_excel('representantes_de_entidades.xlsx', sheet_name='Entidades')

    table = df[
        ['RNC', 'CONTACTO', 'APELLIDO', 'IDENTIFICACION', 'SEXO', 'FECHA NACIMIENTO', 'TELEFONO', 'DIRECCION',
         'NACIONALIDAD', 'OCUPACION']]

    # Loop can do better
    for ind, reg in table.iterrows():
        if reg['RNC'].strip() == key.strip():
            return reg[
                ['CONTACTO', 'APELLIDO', 'IDENTIFICACION', 'SEXO', 'FECHA NACIMIENTO', 'TELEFONO', 'DIRECCION',
                 'NACIONALIDAD', 'OCUPACION']]

def row_generator(obj: XMLFormatter):
    """
    Note on multiprocessing: This is a top-level function, enabling object to be pickled by mp.SimpleQueue
    which is used as default by mp.Pool
    """
    obj.gen_xml()

#? ------------------------------------------ #


if __name__ == '__main__':
    # ? ------------------------ main program top-level 1st version ----------------------------#
    # app = Converter()
    # app.run()
    # print(f"Program execution with timeit module: {timeit.timeit('app', globals=globals())}")
    # ? -----------------------------------------------------------------------------------------#
    start = time()
    c = ConverterFactory().create_obj('xml')
    print(c.name)
    # print(len(c))
    row_generator(c)
    print(f'Execution time: {time() - start}')