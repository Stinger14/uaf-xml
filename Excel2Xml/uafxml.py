"""
    Project: UAF goXML generator app
    Description: Aplicación web de pequeña escala para la generación de archivos XML requeridos por la UAF.
    Dev: Maxly García
"""

from xml.dom import minidom
import xml.etree.ElementTree as ET
import xml.dom.minidom
import openpyxl as xl
import pandas as pd
from copy import deepcopy
from dicttoxml import dicttoxml
import pathlib

# BASE_DIR = pathlib.Path(__file__).resolve().parent
# TEMPLATE = get_resources_path("data/_Web_Report_ReportID_3234-0-0.xml")
# WORKBOOK = get_resources_path("data/transacciones_efectivo_2019_alt.xlsx")
# RTEMAP = get_resources_path("data/mapped_elements.xlsx")

# Reference file to xml tree
TEMPLATE = "_Web_Report_ReportID_3234-0-0.xml"

# RTE excel file.
WORKBOOK = "transacciones_efectivo_2019_copy.xlsx"

# UAF RTE key/value relation - Map UAF fields with bagricola's RTE fields
RTEMAP = "mapped_elements.xlsx"


class RteXml:
    """XML object representation of bagricola's RTE excel file."""

    def __init__(self, template, workbook, rtemap):
        # xml template
        self._template = template
        # excel column names
        self._rtemap = rtemap
        # RTE excel file
        self._workbook = workbook
        # RTE column names
        self.keys = []
        # XML elements
        self.elements = []
        # Root child nodes's child
        self.reports = []

        self.reports2 = []
        # Dictionary made from column names
        self._dict_from_keys = {}
        # Xml data per row
        self._xml_values = ()

        self._keymap = {}

        self._domtree = ''

    # ? Get column names of transacciones_efectivo.xlsx
    def _get_rte_keys(self):
        """
            Return key values from xml mapped elemnents of rte's excel file.
        """
        try:
            df = pd.read_excel(self._rtemap)
            rte_map = df[df['RTE Banke'].notna()]

            try:
                for i, j in rte_map.iterrows():
                    self.keys.append(j.values[1])
                    key = ''.join(j.values[1])

            except TypeError as e:
                print(f'This is not a key: {e}')

        except FileNotFoundError as e:
            print(f'The file {e} does not exist.')

    # ? Get element names from XML template.
    def _get_xml_elements(self):
        """
            Return key values from xml mapped elemnents of rte's excel file.
        """
        try:
            df = pd.read_excel(self._rtemap)
            elements_map = df[df['Elementos UAF'].notna()]
            try:
                for i, j in elements_map.iterrows():
                    self.elements.append(j.values[0])
                    e = ''.join(j.values[0])

            except TypeError as e:
                print(f'This is not a key: {e}')

        except FileNotFoundError as e:
            print(f'The file {e} does not exist.')

    # ? Get all nodes from xml file
    def _get_tree(self):
        return minidom.parse("data/" + self._template.name)

    # ? Generate map between RTE module values and XML elements name.
    def _gen_keymap(self):
        try:
            df = pd.DataFrame([self.elements, self.keys])
            self._keymap = dict.fromkeys(self.elements, '')
            for i in range(len(df.columns)):
                for k in self.elements:
                    if self._keymap[k] == '' and k == df.iloc[0][i]:
                        if df.iloc[1][i] in self.keys and df.iloc[1][i] != 'parent node':
                            self._keymap[k] = df.iloc[1][i]
                            break
                        elif df.iloc[1][i] == 'parent node':
                            break
            # print('Keymap generated.')
        except KeyError as e:
            print(e)

    # #? Print parent & child nodes
    def _print_elements(self):
        tree = ET.parse(self._template)
        ET.XML(self._template)
        root = tree.getroot()
        print([elem.tag for elem in root.iter()])

    def get_rteuaf_dict(self):
        """Iterate excel file columns, extract required columns values
            and update xml file, transactions per excel row.
        """

        try:
            # TODO Create dict mapping elemtents to matching RTE Column name value.

            # Excel workbook to extract data from.
            wb = xl.load_workbook("../data/" + self._workbook.name)
            # self._domtree = minidom.parse(self._template)
            ws = wb.worksheets[0]
            tmp_dict = dict.fromkeys(self.keys, '')

            # ? Generate values mapping
            self._gen_keymap()

            for sheet in wb:
                # ! Bug: Each report is overwritten by the next one.
                for row in range(8, 9):
                    for col in range(2, ws.max_column + 1):
                        col_name = sheet.cell(row=7, column=col).value
                        cellvalue = sheet.cell(row=row, column=col).value
                        if col_name in self.keys:
                            for k in tmp_dict:
                                if col_name == k:
                                    tmp_dict[k] = cellvalue
                                    continue

                    # ! Jump to next row.

                    # TODO Call _update_xml on current dict
                    y = list(self._xml_values)
                    y.append(tmp_dict)
                    self._xml_values = tuple(y)
                    # xmltmp = self._update_xml(tmp_dict)

                    # ? Return dict from RTE
                    # self.reports.append(self._xml_values)
                    # tmp_dict = dict.fromkeys(self.keys, '')
                    # self._dict_from_keys = dict(tmp_dict.keys, tmp_dict.values)
                    print('Report generated.')
                    # return tmp_dict
                    # break

            # self._update_xml(tmp_dict, row)
            self.reports2.append(self._xml_values)
            return tmp_dict
        except KeyError as e:
            print(e)

    # TODO Update xml file.
    def _update_xml(self, adict):
        """Updates goAML XML file.
        """
        try:
            domtree = minidom.parse("../data/" + self._template.name)
            report = domtree.documentElement
            new_cols = ['CODIGO DE SUBMISION', 'CODIGO DE REPORTE', 'ENTITY_REFERENCE', 'FIU_REF_NUMBER	ACTION',
                        'TRANSACTION', 'TITULO DE LA PERSONA', 'FECHA DE NACIMIENTO', 'TELEFONOS INTERMEDIARIO',
                        'EMAIL', 'TELEFONO', 'TPH_CONTACT_TYPE', 'TPH_COMMUNICATION_TYPE', 'TPH_NUMBER', 'TPH_COUNTRY_PREFIX',
                        'TPH_EXTENSION', 'ADDRESS_TYPE', 'CIUDAD', 'COUNTRY_CODE', 'TRANSACTIONNUMBER',
                        'ROL', 'PERSONA CLIENTE', 'COUNTRY', 'ID PERSONA', 'CLIENTE', 'ORIGEN QUEJA',
                        'TELEFONO REPORTANTE', 'FECHA ENVIO', 'MOTIVO DEL REPORTE', 'TRANSMODE_COMMENT', 'ZIP', 'COUNTRY',
                        'COMENTARIOS FONDOS', 'CODIGO FONDOS', 'PUEBLO', 'DIRECCION', 'COMENTARIOS', 'ROL', 'CODIGO DE FONDOS',
                        'COMENTARIO SOBRE FONDOS', 'PAIS DEL CLIENTE', '2do NOMBRE CLIENTE TRANSACCION', 'APELLIDO CLIENTE', 'TRANSACCION', 'SSN',
                        'DIRECCION DE EMPLEADOR', 'TELEFONO DE EMPLEADOR', 'CEDULA CLIENTE', 'FECHA DE ENVIO', 'FECHA DE EXPIRACION',
                        'PAIS DE ORIGEN']

            # ? Root's parent nodes
            reporting_person = report.getElementsByTagName('reporting_person')
            location = report.getElementsByTagName('location')
            transaction = report.getElementsByTagName('transaction')
            transac_tmp = deepcopy(transaction)

            skip_nodes = ['reporting_person', 'location', 'transaction']

            # ? START LOOPING XML FILE.

            for cn in report.childNodes:
                if cn.nodeName in self.elements:
                    if cn.nodeName in skip_nodes or self._keymap[cn.nodeName] in new_cols:
                        continue
                    cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                    print(f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')

                # ? ------------------------- END ROOT -----------------------------

            for rp in reporting_person:
                print("================ reporting_person's child (dynamic loop) ================")
                for cn in rp.childNodes:
                    if cn.nodeName == 'phones':
                        print("================ reporting_person's phones children ================")
                        for phone in cn.getElementsByTagName('phone'):
                            for ph in phone.childNodes:
                                if ph.nodeName in self.elements:
                                    if ph.nodeName in skip_nodes or self._keymap[cn.nodeName] in new_cols:
                                        continue
                                    ph.childNodes[0].data = adict[self._keymap[ph.nodeName]]
                                    print(f'Key updated: {ph.nodeName} -> {ph.childNodes[0].data}')
                    elif cn.nodeName == 'addresses':
                        print("================ reporting_person's addresses children ================")
                        for address in cn.getElementsByTagName('address'):
                            for ad in address.childNodes:
                                if ad.nodeName in self.elements:
                                    if ad.nodeName in skip_nodes or self._keymap[cn.nodeName] in new_cols:
                                        continue
                                    ad.childNodes[0].data = adict[self._keymap[ad.nodeName]]
                                    print(f'Key updated: {ad.nodeName} -> {ad.childNodes[0].data}')
                    elif cn.nodeName in self.elements:
                        if cn.nodeName in skip_nodes or self._keymap[cn.nodeName] in new_cols:
                            continue
                        cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                        print(f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')

            for loc in location:
                print("================ Root's location children ================")
                for l in loc.childNodes:
                    if l.nodeName == '#text':
                        continue
                    elif self._keymap[l.nodeName] in new_cols:
                        continue
                    elif l.nodeName in self.elements:
                        l.childNodes[0].data = adict[self._keymap[l.nodeName]]
                        print(f'Key updated: {l.nodeName} -> {l.childNodes[0].data}')

            # TODO Transaction element
            for transac in transaction:
                print("================ transaction's children ================")
                for tr in transac.childNodes:
                    if tr.nodeName in self.elements or tr.nodeName == '#text':
                        if tr.nodeName == 'involved_parties':
                            print("         ========= transaction's involved_parties children =========      ")
                            for ip in tr.getElementsByTagName('party'):
                                for party in ip.childNodes:
                                    if party.nodeName == 'person_my_client':
                                        print("         ======== transaction's person_my_client children ========")
                                        for pmc in party.childNodes:
                                            if pmc.nodeName in self.elements or pmc.nodeName == '#text':
                                                if pmc.nodeName in skip_nodes:
                                                    continue
                                                if pmc.nodeName == 'phones':
                                                    for p in pmc.getElementsByTagName('phone'):
                                                        print(
                                                            "======== person_my_client's phone children ========")
                                                        for cn in p.childNodes:
                                                            if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                                if cn.nodeName == '#text':
                                                                    continue
                                                                elif self._keymap[cn.nodeName] in new_cols:
                                                                    continue
                                                                cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                                print(
                                                                    f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'addresses':
                                                    for p in pmc.getElementsByTagName('address'):
                                                        print(
                                                            "======== person_my_client's address children ========")
                                                        for cn in p.childNodes:
                                                            if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                                if cn.nodeName == '#text':
                                                                    continue
                                                                elif self._keymap[cn.nodeName] in new_cols:
                                                                    continue
                                                                cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                                print(
                                                                    f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                        break
                                                elif pmc.nodeName == 'employer_address_id':
                                                    print(
                                                        "======== person_my_client's employer_address_id children "
                                                        "=======")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'employer_phone_id':
                                                    print(
                                                        "======== person_my_client's employer_phone_id children "
                                                        "=======")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'identification':
                                                    print(
                                                        "======== person_my_client's identification children ========")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                else:
                                                    # ? All non parent nodes
                                                    if pmc.nodeName == '#text':
                                                        continue
                                                    elif self._keymap[pmc.nodeName] in new_cols:
                                                        continue
                                                    pmc.childNodes[0].data = adict[self._keymap[pmc.nodeName]]
                                                    print(f'Key updated: {pmc.nodeName} -> {pmc.childNodes[0].data}')

                                    elif party.nodeName in self.elements and party.nodeName != 'person_my_client':
                                        if party.nodeName == '#text':
                                            continue
                                        elif self._keymap[party.nodeName] in new_cols:
                                            continue
                                        party.childNodes[0].data = adict[self._keymap[party.nodeName]]
                                        print(f'Key updated: {party.nodeName} -> {party.childNodes[0].data}')

                        # ? If not involved_parties
                        if tr.nodeName != 'involved_parties' and tr.nodeName != '#text':
                            if tr.nodeName == '#text':
                                continue
                            elif self._keymap[tr.nodeName] in new_cols:
                                continue
                            tr.childNodes[0].data = adict[self._keymap[tr.nodeName]]
                            print(f'Key updated: {tr.nodeName} -> {tr.childNodes[0].data}')

                            # root element
            print(domtree.firstChild.tagName)
            domtree.writexml(open('data_report.xml', 'w'))
            # return domtree
            # ! END REPORT

        except Exception as e:
            print(e)

    def update_transac(self, adict):
        try:
            domtree = xml.dom.minidom.parse(self._template)

            # domtree = ET.parse(self._template)
            report = domtree.documentElement
            transaction = report.getElementsByTagName('transaction')
            skip_nodes = ['transaction']

            new_cols = ['CODIGO DE SUBMISION', 'CODIGO DE REPORTE', 'ENTITY_REFERENCE', 'FIU_REF_NUMBER	ACTION', \
                        'TRANSACTION', 'TITULO DE LA PERSONA', 'FECHA DE NACIMIENTO', 'TELEFONOS INTERMEDIARIO',
                        'EMAIL', 'TELEFONO', 'TPH_CONTACT_TYPE', 'TPH_COMMUNICATION_TYPE', 'TPH_NUMBER', 'TPH_COUNTRY_PREFIX', \
                        'TPH_EXTENSION', 'ADDRESS_TYPE', 'CIUDAD', 'COUNTRY_CODE', 'TRANSACTIONNUMBER', \
                        'ROL', 'PERSONA CLIENTE', 'COUNTRY', 'ID PERSONA', 'CLIENTE', 'ORIGEN QUEJA',
                        'TELEFONO REPORTANTE', 'FECHA ENVIO', 'MOTIVO DEL REPORTE', 'TRANSMODE_COMMENT', 'ZIP', 'COUNTRY',
                        'COMENTARIOS FONDOS', 'CODIGO FONDOS', 'PUEBLO', 'DIRECCION', 'COMENTARIOS', 'ROL', 'CODIGO DE FONDOS',
                        'COMENTARIO SOBRE FONDOS', 'PAIS DEL CLIENTE', '2do NOMBRE CLIENTE TRANSACCION', 'APELLIDO CLIENTE', 'TRANSACCION', 'SSN',
                        'DIRECCION DE EMPLEADOR', 'TELEFONO DE EMPLEADOR', 'CEDULA CLIENTE', 'FECHA DE ENVIO', 'FECHA DE EXPIRACION',
                        'PAIS DE ORIGEN']

            # length = len(adict)
            for transac in transaction:
                print("================ transaction's children ================")
                for tr in transac.childNodes:
                    if tr.nodeName in self.elements or tr.nodeName == '#text':
                        if tr.nodeName == 'involved_parties':
                            print("         ========= transaction's involved_parties children =========      ")
                            for ip in tr.getElementsByTagName('party'):
                                for party in ip.childNodes:
                                    if party.nodeName == 'person_my_client':
                                        print("         ======== transaction's person_my_client children ========")
                                        for pmc in party.childNodes:
                                            if pmc.nodeName in self.elements or pmc.nodeName == '#text':
                                                if pmc.nodeName in skip_nodes:
                                                    continue
                                                if pmc.nodeName == 'phones':
                                                    for p in pmc.getElementsByTagName('phone'):
                                                        print(
                                                            "         ======== person_my_client's phone children ========")
                                                        for cn in p.childNodes:
                                                            if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                                if cn.nodeName == '#text':
                                                                    continue
                                                                elif self._keymap[cn.nodeName] in new_cols:
                                                                    continue
                                                                cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                                print(
                                                                    f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'addresses':
                                                    for p in pmc.getElementsByTagName('address'):
                                                        print(
                                                            "======== person_my_client's address children ========")
                                                        for cn in p.childNodes:
                                                            if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                                if cn.nodeName == '#text':
                                                                    continue
                                                                elif self._keymap[cn.nodeName] in new_cols:
                                                                    continue
                                                                cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                                print(
                                                                    f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                        break
                                                elif pmc.nodeName == 'employer_address_id':
                                                    print(
                                                        "======== person_my_client's employer_address_id children "
                                                        "=======")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'employer_phone_id':
                                                    print(
                                                        "======== person_my_client's employer_phone_id children "
                                                        "=======")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                elif pmc.nodeName == 'identification':
                                                    print(
                                                        "======== person_my_client's identification children ========")
                                                    for cn in pmc.childNodes:
                                                        if cn.nodeName in self.elements or cn.nodeName == '#text':
                                                            if cn.nodeName == '#text':
                                                                continue
                                                            if self._keymap[cn.nodeName] in new_cols:
                                                                continue
                                                            cn.childNodes[0].data = adict[self._keymap[cn.nodeName]]
                                                            print(
                                                                f'Key updated: {cn.nodeName} -> {cn.childNodes[0].data}')
                                                else:
                                                    # ? All non parent nodes
                                                    if cn.nodeName == '#text':
                                                        continue
                                                    elif self._keymap[cn.nodeName] in new_cols:
                                                        continue
                                                    pmc.childNodes[0].data = adict[self._keymap[pmc.nodeName]]
                                                    print(f'Key updated: {pmc.nodeName} -> {pmc.childNodes[0].data}')

                                    elif party.nodeName in self.elements and party.nodeName != 'person_my_client':
                                        if party.nodeName == '#text':
                                            continue
                                        elif self._keymap[party.nodeName] in new_cols:
                                            continue
                                        party.childNodes[0].data = adict[self._keymap[party.nodeName]]
                                        print(f'Key updated: {party.nodeName} -> {party.childNodes[0].data}')

                        # ? If not involved_parties
                        if tr.nodeName != 'involved_parties' and tr.nodeName != '#text':
                            if tr.nodeName == '#text':
                                continue
                            elif self._keymap[tr.nodeName] in new_cols:
                                continue
                            tr.childNodes[0].data = adict[self._keymap[tr.nodeName]]
                            print(f'Key updated: {tr.nodeName} -> {tr.childNodes[0].data}')
            domtree.writexml(open('data_report.xml', 'w'))
            # ! END REPORT
        except KeyError as e:
            print(e)

# if __name__ == "__main__":
#     #TODO 1) Generate tree from xml file
#     xml_template = RteXml(TEMPLATE, WORKBOOK, RTEMAP)
#     # tree = xml_template._get_tree()
#     # print(f'Tree: {tree}')

#     #TODO 2) Create dict with RTE Excel column names
#     xml_template._get_rte_keys()
#     # print(xml_template.keys)

#     #TODO 3) Create dict with XML element names
#     xml_template._get_xml_elements()
#     # print(xml_template.elements)

#     # TODO 4) Update xml file with dict updated values
#     adict = xml_template.get_rteuaf_dict()
#     print(adict)
#     print(f'Número de transacciones: {len(adict)}')

#     #TODO Print xml elements
#     # xml_template._print_elements()

#     # TODO Update xml file properties
#     # xml_template.update_transac(adict)
#     # print('Xml tree updated.')
